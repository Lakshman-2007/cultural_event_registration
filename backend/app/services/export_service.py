"""Export service for generating CSV and Excel exports of participant data."""

import csv
import io
from typing import Optional, List
from openpyxl import Workbook
from sqlalchemy.orm import Session
from sqlalchemy import or_

from app.models import Participant


# Column headers for export
EXPORT_COLUMNS = [
    "Registration ID",
    "Full Name",
    "Mobile",
    "Email",
    "Address",
    "College Name",
    "Register Number",
    "Department",
    "Year of Study",
    "Participant Type",
    "Payment Status",
    "Payment Amount",
    "Transaction ID",
    "Attendance Status",
    "Checked In At",
    "Created At",
]


def _get_filtered_participants(
    db: Session,
    search: Optional[str] = None,
    college: Optional[str] = None,
    payment_status: Optional[str] = None,
    attendance_status: Optional[str] = None,
) -> List[Participant]:
    """
    Get filtered list of participants for export.
    
    Args:
        db: Database session
        search: Search term for name, register number, or mobile
        college: Filter by college name
        payment_status: Filter by payment status
        attendance_status: Filter by attendance status
        
    Returns:
        List of Participant objects matching the filters
    """
    query = db.query(Participant)

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Participant.full_name.ilike(search_term),
                Participant.register_number.ilike(search_term),
                Participant.mobile.ilike(search_term),
                Participant.registration_id.ilike(search_term),
                Participant.email.ilike(search_term),
            )
        )

    if college:
        query = query.filter(Participant.college_name.ilike(f"%{college}%"))

    if payment_status:
        query = query.filter(Participant.payment_status == payment_status.upper())

    if attendance_status:
        query = query.filter(Participant.attendance_status == attendance_status.upper())

    return query.order_by(Participant.created_at.desc()).all()


def _participant_to_row(participant: Participant) -> list:
    """Convert a participant model to an export row."""
    return [
        participant.registration_id,
        participant.full_name,
        participant.mobile,
        participant.email,
        participant.address,
        participant.college_name,
        participant.register_number,
        participant.department,
        participant.year_of_study,
        participant.participant_type,
        participant.payment_status,
        participant.payment_amount,
        participant.transaction_id or "",
        participant.attendance_status,
        participant.checked_in_at.isoformat() if participant.checked_in_at else "",
        participant.created_at.isoformat() if participant.created_at else "",
    ]


def generate_csv_export(
    db: Session,
    search: Optional[str] = None,
    college: Optional[str] = None,
    payment_status: Optional[str] = None,
    attendance_status: Optional[str] = None,
) -> str:
    """
    Generate a CSV string of participant data with optional filtering.
    
    Returns:
        CSV formatted string
    """
    participants = _get_filtered_participants(db, search, college, payment_status, attendance_status)

    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(EXPORT_COLUMNS)

    # Write data rows
    for participant in participants:
        writer.writerow(_participant_to_row(participant))

    return output.getvalue()


def generate_excel_export(
    db: Session,
    search: Optional[str] = None,
    college: Optional[str] = None,
    payment_status: Optional[str] = None,
    attendance_status: Optional[str] = None,
) -> bytes:
    """
    Generate an Excel (xlsx) file of participant data with optional filtering.
    
    Returns:
        Bytes of the Excel file
    """
    participants = _get_filtered_participants(db, search, college, payment_status, attendance_status)

    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    # Write header with formatting
    for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)

    # Write data rows
    for row_idx, participant in enumerate(participants, 2):
        row_data = _participant_to_row(participant)
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
        max_length = len(header)
        for row_idx in range(2, len(participants) + 2):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
